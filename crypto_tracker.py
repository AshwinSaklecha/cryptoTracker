import requests
import pandas as pd
import time
from datetime import datetime
import os
import openpyxl

class CryptoTracker:
    def __init__(self):
        self.base_url = "https://api.coingecko.com/api/v3"
        self.excel_file = "crypto_data.xlsx"
        
    def fetch_top_50_crypto(self):
        """Fetch top 50 cryptocurrencies data from CoinGecko"""
        try:
            endpoint = f"{self.base_url}/coins/markets"
            params = {
                'vs_currency': 'usd',
                'order': 'market_cap_desc',
                'per_page': 50,
                'page': 1,
                'sparkline': False
            }
            
            response = requests.get(endpoint, params=params)
            response.raise_for_status()
            return response.json()
            
        except requests.RequestException as e:
            print(f"Error fetching data: {e}")
            return None

    def process_crypto_data(self, data):
        """Process the raw cryptocurrency data into a pandas DataFrame"""
        if not data:
            return None
            
        df = pd.DataFrame(data)
        df = df[[
            'name',
            'symbol',
            'current_price',
            'market_cap',
            'total_volume',
            'price_change_percentage_24h'
        ]]
        
        df.columns = [
            'Cryptocurrency Name',
            'Symbol',
            'Current Price (USD)',
            'Market Capitalization',
            '24h Trading Volume',
            'Price Change 24h (%)'
        ]
        
        return df

    def analyze_data(self, df):
        """Perform analysis on the cryptocurrency data"""
        if df is None or df.empty:
            return None
            
        analysis = {
            'Top 5 by Market Cap': df.head(5)[['Cryptocurrency Name', 'Market Capitalization']],
            'Average Price': df['Current Price (USD)'].mean(),
            'Highest 24h Change': df.nlargest(1, 'Price Change 24h (%)')[['Cryptocurrency Name', 'Price Change 24h (%)']],
            'Lowest 24h Change': df.nsmallest(1, 'Price Change 24h (%)')[['Cryptocurrency Name', 'Price Change 24h (%)']],
        }
        
        return analysis

    def update_excel(self, df, analysis):
        """Update Excel file with latest data and analysis"""
        try:
            # Prepare the data - convert symbols to uppercase
            df['Symbol'] = df['Symbol'].str.upper()
            
            with pd.ExcelWriter(self.excel_file, engine='openpyxl', mode='w') as writer:
                # Write main data
                df.to_excel(writer, sheet_name='Live Crypto Data', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Live Crypto Data']
                
                # Define styles
                header_style = {
                    'fill': {'fgColor': '366092', 'patternType': 'solid'},
                    'font': {'color': 'FFFFFF', 'bold': True},
                    'border': {'style': 'thin', 'color': '000000'}
                }
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
                # Apply header styling
                for cell in worksheet[1]:
                    cell.fill = openpyxl.styles.PatternFill(start_color=header_style['fill']['fgColor'],
                                                           end_color=header_style['fill']['fgColor'],
                                                           fill_type=header_style['fill']['patternType'])
                    cell.font = openpyxl.styles.Font(color=header_style['font']['color'],
                                                   bold=header_style['font']['bold'])
                    cell.border = openpyxl.styles.Border(top=openpyxl.styles.Side(style='thin'),
                                                       bottom=openpyxl.styles.Side(style='thin'),
                                                       left=openpyxl.styles.Side(style='thin'),
                                                       right=openpyxl.styles.Side(style='thin'))
                
                # Format numbers
                price_format =  '#,##0.00' # Currency with 2 decimal places
                large_number_format = '#,##0'
                percent_format = '0.00%'
                
                # Apply number formatting and conditional formatting for price changes
                for row in worksheet.iter_rows(min_row=2):  # Skip header
                    # Price formatting
                    row[2].number_format = price_format  # Current Price
                    row[3].number_format = large_number_format  # Market Cap
                    row[4].number_format = large_number_format  # Trading Volume
                    
                    # Price change formatting and colors
                    price_change_cell = row[5]
                    price_change_cell.number_format = percent_format
                    if price_change_cell.value:
                        if float(price_change_cell.value) > 0:
                            price_change_cell.fill = openpyxl.styles.PatternFill(
                                start_color='C6EFCE',
                                end_color='C6EFCE',
                                fill_type='solid'
                            )
                            price_change_cell.font = openpyxl.styles.Font(color='006100')
                        elif float(price_change_cell.value) < 0:
                            price_change_cell.fill = openpyxl.styles.PatternFill(
                                start_color='FFC7CE',
                                end_color='FFC7CE',
                                fill_type='solid'
                            )
                            price_change_cell.font = openpyxl.styles.Font(color='9C0006')
                
                # Write analysis sheets with similar formatting
                analysis['Top 5 by Market Cap'].to_excel(writer, sheet_name='Analysis', startrow=1, index=False)
                pd.DataFrame({'Average Price (USD)': [analysis['Average Price']]}).to_excel(
                    writer, sheet_name='Analysis', startrow=8, index=False)
                analysis['Highest 24h Change'].to_excel(writer, sheet_name='Analysis', startrow=11, index=False)
                analysis['Lowest 24h Change'].to_excel(writer, sheet_name='Analysis', startrow=14, index=False)
                
            print(f"Excel file updated at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
        except Exception as e:
            print(f"Error updating Excel file: {e}")

    def run(self, update_interval=300):
        """Run the crypto tracker with specified update interval (in seconds)"""
        while True:
            # Fetch and process data
            raw_data = self.fetch_top_50_crypto()
            df = self.process_crypto_data(raw_data)
            
            if df is not None:
                # Perform analysis
                analysis = self.analyze_data(df)
                
                # Update Excel
                self.update_excel(df, analysis)
                
            # Wait for the specified interval
            time.sleep(update_interval)

if __name__ == "__main__":
    tracker = CryptoTracker()
    # Update every 5 minutes (300 seconds)
    tracker.run(update_interval=300) 